import streamlit as st
import pandas as pd
from datetime import date
import hashlib
import io

# ──────────────────────────────────────────────
# 페이지 기본 설정
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="요양병원 간호관리료 등급 산정",
    page_icon="🏥",
    layout="wide",
)

# ──────────────────────────────────────────────
# CSS
# ──────────────────────────────────────────────
st.markdown("""
<style>
    .main-title {
        font-size: 24px; font-weight: 800; color: #0d4f3c;
        border-bottom: 3px solid #0d4f3c; padding-bottom: 10px; margin-bottom: 16px;
        display: flex; align-items: baseline; gap: 12px;
    }
    .creator-badge { font-size: 13px; color: #888; font-weight: 500; }
    .section-title {
        font-size: 15px; font-weight: 700; color: #0d4f3c;
        background: #eaf6f0; border-left: 5px solid #0d4f3c;
        padding: 7px 12px; margin: 14px 0 8px 0; border-radius: 0 6px 6px 0;
    }
    .sub-title {
        font-size: 13px; font-weight: 700; color: #0d4f3c;
        margin: 10px 0 6px 0;
    }
    .result-card {
        background: #f0fbf5; border: 1.5px solid #a6ddc0;
        border-radius: 10px; padding: 14px 20px; margin: 8px 0;
    }
    .grade-box {
        display: inline-block; font-size: 30px; font-weight: 900;
        padding: 10px 26px; border-radius: 12px; color: white; margin: 4px 6px;
    }
    .grade-n1 { background: #0d47a1; }
    .grade-n2 { background: #1976d2; }
    .grade-n3 { background: #2e7d32; }
    .grade-n4 { background: #f57f17; }
    .grade-n5 { background: #ef6c00; }
    .grade-n6 { background: #b71c1c; }
    .grade-d1 { background: #4527a0; }
    .grade-d2 { background: #00695c; }
    .grade-d3 { background: #6d4c41; }
    .grade-d4 { background: #b71c1c; }
    .kpi-label { font-size: 12px; color: #555; margin-bottom: 2px; }
    .kpi-value { font-size: 20px; font-weight: 700; color: #0d4f3c; }
    .kpi-unit  { font-size: 11px; color: #777; }
    .yellow-note {
        background: #fffde7; border: 1px solid #f9a825;
        border-radius: 6px; padding: 7px 12px; font-size: 12px; color: #5d4037;
        margin-bottom: 6px;
    }
    .blue-note {
        background: #e8f3ff; border: 1px solid #90b8e0;
        border-radius: 6px; padding: 7px 12px; font-size: 12px; color: #1a3a6b;
        margin-bottom: 6px;
    }
    .footer {
        font-size: 13px; color: #555; text-align: center;
        margin-top: 30px; border-top: 1px solid #ddd; padding-top: 12px;
    }
    @media print {
        header, footer,
        [data-testid="stToolbar"], [data-testid="stSidebar"],
        [data-testid="stDecoration"], [data-testid="stStatusWidget"],
        .stButton > button, .stDownloadButton { display: none !important; }
        @page { size: A4 landscape; margin: 8mm; }
        html, body { margin: 0 !important; padding: 0 !important; }
        [data-testid="block-container"] {
            padding: 6px 10px !important; max-width: 100% !important; width: 100% !important;
        }
        * { font-size: 9px !important; line-height: 1.3 !important; }
        .main-title  { font-size: 13px !important; }
        .section-title { font-size: 10px !important; }
        .kpi-value { font-size: 12px !important; }
        .grade-box { font-size: 16px !important; padding: 5px 12px !important; }
        table, th, td { font-size: 9px !important; padding: 2px 4px !important; }
        .footer { margin-top: 10px !important; padding-top: 6px !important; }
    }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# 분기 / 날짜 유틸
# ──────────────────────────────────────────────
QUARTER_RANGES = {
    "1분기 (12/15 ~ 3/14)": {"month_start": 12, "day_start": 15, "month_end": 3,  "day_end": 14},
    "2분기 (3/15 ~ 6/14)":  {"month_start": 3,  "day_start": 15, "month_end": 6,  "day_end": 14},
    "3분기 (6/15 ~ 9/14)":  {"month_start": 6,  "day_start": 15, "month_end": 9,  "day_end": 14},
    "4분기 (9/15 ~ 12/14)": {"month_start": 9,  "day_start": 15, "month_end": 12, "day_end": 14},
}

def get_quarter_dates(year, quarter_label):
    q = QUARTER_RANGES[quarter_label]
    if quarter_label.startswith("1"):
        start = date(year - 1, q["month_start"], q["day_start"])
        end   = date(year,     q["month_end"],   q["day_end"])
    else:
        start = date(year, q["month_start"], q["day_start"])
        end   = date(year, q["month_end"],   q["day_end"])
    return start, end, (end - start).days + 1

def calc_active_days(hire_date, status, q_start, q_end, resign_date=None):
    if hire_date is None or hire_date > q_end:
        return 0
    if status == "퇴사":
        if resign_date is None or resign_date < q_start:
            return 0
        effective_end = min(resign_date, q_end)
    else:
        effective_end = q_end
    effective_start = max(hire_date, q_start)
    return max(0, (effective_end - effective_start).days + 1)

def month_label(base, offset):
    m = ((base.month - 1 + offset) % 12) + 1
    y = base.year + ((base.month - 1 + offset) // 12)
    return f"{y}년 {m}월"

# 요양병원 간호인력 단시간 근무 가중치 (일반병동과 다른 8단계 체계)
PARTTIME_TABLE_NORMAL = {
    "전일제 (주40h 이상)":        1.0,
    "단시간 (주36~40h 미만)":     0.9,
    "단시간 (주32~36h 미만)":     0.8,
    "단시간 (주28~32h 미만)":     0.7,
    "단시간 (주24~28h 미만)":     0.6,
    "단시간 (주20~24h 미만)":     0.5,
    "단시간 (주16~20h 미만)":     0.4,
}
PARTTIME_TABLE_REMOTE = {  # 의료취약지역 소재 요양기관
    "전일제 (주40h 이상)":        1.0,
    "단시간 (주36~40h 미만)":     1.0,
    "단시간 (주32~36h 미만)":     0.9,
    "단시간 (주28~32h 미만)":     0.8,
    "단시간 (주24~28h 미만)":     0.7,
    "단시간 (주20~24h 미만)":     0.6,
    "단시간 (주16~20h 미만)":     0.5,
}
WORKTYPE_OPTS = list(PARTTIME_TABLE_NORMAL.keys())

PAYER_COLS = ["건강보험", "의료급여", "자보", "산재", "기타"]
PAYER_KEYS = ["hi", "mc", "auto", "labor", "etc"]

# ──────────────────────────────────────────────
# 등급 산정 로직
# ──────────────────────────────────────────────
def determine_nursing_grade(ratio):
    """환자수 대 간호인력수(간호사+간호조무사) 비율 → 1~6등급"""
    if   ratio < 4.5: return 1
    elif ratio < 5.0: return 2
    elif ratio < 5.5: return 3
    elif ratio < 6.0: return 4
    elif ratio < 6.5: return 5
    else:             return 6

NURSING_BONUS_PCT = {1: 60, 2: 50, 3: 35, 4: 20, 5: 0, 6: -50}
NURSING_GRADE_TABLE = [
    (1, "4.5:1 미만",        "+60%"),
    (2, "4.5:1 이상 5.0:1 미만", "+50%"),
    (3, "5.0:1 이상 5.5:1 미만", "+35%"),
    (4, "5.5:1 이상 6.0:1 미만", "+20%"),
    (5, "6.0:1 이상 6.5:1 미만", "0%"),
    (6, "6.5:1 이상",        "-50%"),
]

def determine_doctor_grade(ratio, specialist_ratio):
    """환자수 대 의사수 비율 + 전문의 비율 → 1~4등급"""
    if ratio <= 35:
        return 1 if specialist_ratio >= 0.5 else 2
    elif ratio <= 40:
        return 3
    else:
        return 4

DOCTOR_BONUS_PCT = {1: 13, 2: 5, 3: 0, 4: -50}
DOCTOR_GRADE_TABLE = [
    (1, "35:1 이하  · 전문의 비율 50% 이상", "+13%"),
    (2, "35:1 이하  · 전문의 비율 50% 미만", "+5%"),
    (3, "35:1 초과 40:1 이하",             "0%"),
    (4, "40:1 초과",                       "-50%"),
]

def grade_css_n(g): return f"grade-n{g}"
def grade_css_d(g): return f"grade-d{g}"

def next_better_nursing_grade(g):
    return g - 1 if g > 1 else None

def next_better_doctor_grade(g):
    return g - 1 if g > 1 else None

# ──────────────────────────────────────────────
# 엑셀 업로드 파싱
# ──────────────────────────────────────────────
def parse_excel_upload(file_bytes):
    import openpyxl
    from datetime import date as date_type, datetime as datetime_type

    def parse_date(v):
        if v is None: return None
        if isinstance(v, datetime_type): return v.date()
        if isinstance(v, date_type): return v
        s = str(v).strip()
        if not s: return None
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d",
                    "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
            try:
                from datetime import datetime as dt
                return dt.strptime(s, fmt).date()
            except Exception:
                pass
        return None

    def to_int(v):
        if v is None: return 0
        try: return int(float(str(v)))
        except Exception: return 0

    def to_str(v):
        return str(v).strip() if v is not None else None

    def to_bool_yn(v, default=False):
        s = to_str(v)
        if s is None: return default
        return s.upper() in ("Y", "YES", "예", "O")

    result = {
        "hosp_name": "", "year": 2026, "quarter": None, "remote_area": False,
        "patients": {i: {k: 0 for k in PAYER_KEYS} for i in range(3)},
        "daycare": {i: {k: 0 for k in PAYER_KEYS} for i in range(3)},
        "nurses": [], "aides": [], "doctors": [], "necessary": {},
    }
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        if "기본정보" in wb.sheetnames:
            ws = wb["기본정보"]
            result["hosp_name"] = to_str(ws["B6"].value) or ""
            y = to_int(ws["D6"].value)
            result["year"] = y if y else 2026
            result["quarter"] = to_str(ws["F6"].value)
            result["remote_area"] = to_bool_yn(ws["H6"].value, False)

        if "환자수현황" in wb.sheetnames:
            ws2 = wb["환자수현황"]
            for i in range(3):
                in_row = 5 + i * 2
                day_row = 6 + i * 2
                for j, k in enumerate(PAYER_KEYS):
                    col = 3 + j  # C,D,E,F,G
                    result["patients"][i][k] = to_int(ws2.cell(in_row, col).value)
                    result["daycare"][i][k] = to_int(ws2.cell(day_row, col).value)

        def read_personnel(sheet_name, extra_cols):
            rows = []
            if sheet_name not in wb.sheetnames: return rows
            ws_p = wb[sheet_name]
            empty = 0
            for r in range(5, 205):
                hr = ws_p.cell(r, 2).value
                if hr is None:
                    empty += 1
                    if empty >= 3: break
                    continue
                empty = 0
                hire = parse_date(hr)
                resign = parse_date(ws_p.cell(r, 3).value)
                status = to_str(ws_p.cell(r, 4).value) or "근무"
                if status not in ["근무", "퇴사"]: status = "근무"
                extra = {name: ws_p.cell(r, col_idx).value for name, col_idx in extra_cols.items()}
                if hire:
                    rows.append({"hire_date": hire, "resign_date": resign if status == "퇴사" else None,
                                 "status": status, **extra})
            return rows

        for row in read_personnel("간호사", {"worktype_raw": 5}):
            wt = to_str(row.pop("worktype_raw"))
            row["worktype"] = wt if wt in WORKTYPE_OPTS else WORKTYPE_OPTS[0]
            result["nurses"].append(row)

        for row in read_personnel("간호조무사", {"worktype_raw": 5, "employ_raw": 6}):
            wt = to_str(row.pop("worktype_raw"))
            row["worktype"] = wt if wt in WORKTYPE_OPTS else WORKTYPE_OPTS[0]
            employ = to_str(row.pop("employ_raw"))
            row["employ"] = employ if employ in ["정규직", "계약직"] else "정규직"
            result["aides"].append(row)

        for row in read_personnel("의사", {"worktype_raw": 5, "specialist_raw": 6}):
            wt = to_str(row.pop("worktype_raw"))
            row["worktype"] = "전일제" if (wt is None or "전일" in wt) else "시간제/격일제(0.5인)"
            row["specialist"] = to_bool_yn(row.pop("specialist_raw"), True)
            result["doctors"].append(row)

        if "필요인력" in wb.sheetnames:
            ws6 = wb["필요인력"]
            labels_map = {
                "약사 상근 여부": "pharm_present",
                "약사 주16시간 이상 근무 (환자 200명 미만 시 인정)": "pharm_hours16",
                "보건의료정보관리사 상근 1인 이상": "hima_present",
                "방사선사 상근 1인 이상": "radio_present",
                "임상병리사 상근 1인 이상": "lab_present",
                "물리치료사 상근 1인 이상": "pt_present",
                "사회복지사 상근 1인 이상": "sw_present",
            }
            for r in range(5, 12):
                label = to_str(ws6.cell(r, 1).value)
                if label in labels_map:
                    result["necessary"][labels_map[label]] = to_bool_yn(ws6.cell(r, 2).value, False)

    except Exception as e:
        return None, str(e)
    return result, None

# ──────────────────────────────────────────────
# 세션 상태 초기화
# ──────────────────────────────────────────────
def init_session():
    defaults = {
        "hosp_name": "", "year": 2026, "quarter_idx": 1, "remote_area": False,
        "nurse_rows": [{"hire_date": None, "resign_date": None, "status": "근무",
                         "worktype": WORKTYPE_OPTS[0]}],
        "aide_rows": [{"hire_date": None, "resign_date": None, "status": "근무",
                        "worktype": WORKTYPE_OPTS[0], "employ": "정규직"}],
        "doctor_rows": [{"hire_date": None, "resign_date": None, "status": "근무",
                          "worktype": "전일제", "specialist": True}],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
    for i in range(3):
        for payer in ["hi", "mc", "auto", "labor", "etc"]:
            key = f"pat_{i}_{payer}"
            if key not in st.session_state:
                st.session_state[key] = 0
            key2 = f"day_{i}_{payer}"
            if key2 not in st.session_state:
                st.session_state[key2] = 0
    for k, v in {
        "pharm_present": True, "pharm_hours16": False,
        "hima_present": True, "radio_present": True,
        "lab_present": True, "pt_present": True, "sw_present": True,
    }.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_session()

# ──────────────────────────────────────────────
# 헤더
# ──────────────────────────────────────────────
st.markdown(
    '<div class="main-title">🏥 요양병원 입원료 차등제 등급 산정 시스템'
    '<span class="creator-badge">ㅣ 제작: 조정윤 · 주식회사 메디엄</span></div>',
    unsafe_allow_html=True
)
st.markdown(
    '<div class="blue-note">📌 본 시스템은 「건강보험 행위 급여·비급여 목록표 및 급여 상대가치점수」 제3편 요양병원 '
    '및 「요양급여의 적용기준 및 방법에 관한 세부사항」의 간호인력·의사인력 확보수준에 따른 입원료 차등제, '
    '필요인력 확보에 따른 별도 보상제 기준을 반영합니다. 고시는 주기적으로 개정되므로 신고 전 반드시 최신 고시 내용을 '
    '심사평가원 홈페이지에서 확인하시기 바랍니다.</div>', unsafe_allow_html=True
)

# ──────────────────────────────────────────────
# 엑셀 입력양식 템플릿 생성 (앱 내부에서 직접 생성 → 별도 파일 불필요)
# ──────────────────────────────────────────────
@st.cache_data
def generate_template_bytes():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as date_type

    FONT_NAME = "Arial"
    HEADER_FILL = PatternFill("solid", fgColor="1A3A6B")
    HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
    INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
    LABEL_FONT = Font(name=FONT_NAME, bold=True, size=10)
    NORMAL_FONT = Font(name=FONT_NAME, size=10)
    TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="0D4F3C")
    NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="B71C1C")
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb_t = openpyxl.Workbook()

    def style_header(ws, cells):
        for c in cells:
            ws[c].fill = HEADER_FILL
            ws[c].font = HEADER_FONT
            ws[c].alignment = Alignment(horizontal="center", vertical="center")
            ws[c].border = BORDER

    def style_input(ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = INPUT_FILL
                cell.font = NORMAL_FONT
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center")

    def set_widths(ws, widths):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    # 시트1: 기본정보
    ws = wb_t.active
    ws.title = "기본정보"
    ws["A1"] = "요양병원 간호관리료 등급산정 - 기본정보 입력"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A3"] = "※ 노란색 셀만 입력하세요. 분기는 아래 목록 중 정확히 동일한 문구로 입력해야 합니다."
    ws["A3"].font = NOTE_FONT
    ws.merge_cells("A3:H3")
    ws["A4"] = "  (1분기 (12/15 ~ 3/14) / 2분기 (3/15 ~ 6/14) / 3분기 (6/15 ~ 9/14) / 4분기 (9/15 ~ 12/14))"
    ws["A4"].font = NOTE_FONT
    ws.merge_cells("A4:H4")
    style_header(ws, ["B5", "D5", "F5", "H5"])
    ws["B5"] = "요양기관명"; ws["D5"] = "연도"; ws["F5"] = "분기"; ws["H5"] = "의료취약지역(Y/N)"
    ws["B6"] = "예시요양병원"; ws["D6"] = 2026; ws["F6"] = "2분기 (3/15 ~ 6/14)"; ws["H6"] = "N"
    style_input(ws, "B6:H6")
    set_widths(ws, {"A": 3, "B": 16, "C": 3, "D": 10, "E": 3, "F": 20, "G": 3, "H": 20})

    # 시트2: 환자수현황
    ws2 = wb_t.create_sheet("환자수현황")
    ws2["A1"] = "월별 재원환자수 · 낮병동환자수 (재원형태별)"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:G1")
    ws2["A3"] = "※ 노란색 셀에 각 월의 구분(입원환자수/낮병동환자수)별, 재원형태별 인원수를 입력하세요."
    ws2["A3"].font = NOTE_FONT
    ws2.merge_cells("A3:G3")
    headers = ["구분", "월", "건강보험", "의료급여", "자보", "산재", "기타"]
    for i, h in enumerate(headers):
        ws2[f"{get_column_letter(1 + i)}4"] = h
    style_header(ws2, ["A4", "B4", "C4", "D4", "E4", "F4", "G4"])
    rows_data = [
        ("입원환자수", "1월", 5067, 1901, 124, 37, 3),
        ("낮병동환자수", "1월", 0, 0, 0, 0, 0),
        ("입원환자수", "2월", 5294, 1828, 134, 57, 0),
        ("낮병동환자수", "2월", 0, 0, 0, 0, 0),
        ("입원환자수", "3월", 4784, 1673, 128, 56, 0),
        ("낮병동환자수", "3월", 0, 0, 0, 0, 0),
    ]
    for r, row in enumerate(rows_data, start=5):
        ws2[f"A{r}"] = row[0]; ws2[f"A{r}"].font = LABEL_FONT
        ws2[f"B{r}"] = row[1]; ws2[f"B{r}"].font = LABEL_FONT
        for j, v in enumerate(row[2:]):
            ws2[f"{get_column_letter(3 + j)}{r}"] = v
        style_input(ws2, f"C{r}:G{r}")
        ws2[f"A{r}"].border = BORDER; ws2[f"B{r}"].border = BORDER
        ws2[f"A{r}"].alignment = Alignment(horizontal="center")
        ws2[f"B{r}"].alignment = Alignment(horizontal="center")
    set_widths(ws2, {"A": 14, "B": 8, "C": 11, "D": 11, "E": 9, "F": 9, "G": 9})

    # 시트3~5: 간호사 / 간호조무사 / 의사
    def build_personnel_sheet(name, extra_headers, extra_widths, example_rows):
        ws_p = wb_t.create_sheet(name)
        ws_p["A1"] = f"{name} 인력 현황 (근무 중 · 최근 퇴사자 모두 입력)"
        ws_p["A1"].font = TITLE_FONT
        ws_p.merge_cells(f"A1:{get_column_letter(4+len(extra_headers))}1")
        ws_p["A3"] = "※ 상태는 '근무' 또는 '퇴사'만 입력. 퇴사가 아니면 퇴사일은 비워두세요."
        ws_p["A3"].font = NOTE_FONT
        ws_p.merge_cells(f"A3:{get_column_letter(4+len(extra_headers))}3")
        headers2 = ["#", "입사일(YYYY-MM-DD)", "퇴사일(YYYY-MM-DD)", "상태(근무/퇴사)"] + extra_headers
        for i, h in enumerate(headers2):
            ws_p[f"{get_column_letter(1 + i)}4"] = h
        style_header(ws_p, [get_column_letter(i + 1) + "4" for i in range(len(headers2))])
        for r_idx, ex in enumerate(example_rows, start=5):
            ws_p[f"A{r_idx}"] = r_idx - 4
            ws_p[f"A{r_idx}"].font = LABEL_FONT
            ws_p[f"A{r_idx}"].alignment = Alignment(horizontal="center")
            ws_p[f"A{r_idx}"].border = BORDER
            for j, v in enumerate(ex):
                ws_p[f"{get_column_letter(2 + j)}{r_idx}"] = v
            last_col = get_column_letter(1 + len(headers2))
            style_input(ws_p, f"B{r_idx}:{last_col}{r_idx}")
        widths = {"A": 4, "B": 16, "C": 16, "D": 12}
        widths.update(extra_widths)
        set_widths(ws_p, widths)

    build_personnel_sheet(
        "간호사", ["근무형태"], {"E": 24},
        [[date_type(2024, 3, 1), None, "근무", WORKTYPE_OPTS[0]],
         [date_type(2025, 6, 1), None, "근무", WORKTYPE_OPTS[1]]],
    )
    build_personnel_sheet(
        "간호조무사", ["근무형태", "고용형태(정규직/계약직)"], {"E": 24, "F": 18},
        [[date_type(2023, 1, 10), None, "근무", WORKTYPE_OPTS[0], "정규직"],
         [date_type(2025, 9, 1), None, "근무", WORKTYPE_OPTS[0], "계약직"]],
    )
    build_personnel_sheet(
        "의사", ["근무형태(전일제/시간제)", "전문의여부(Y/N)"], {"E": 22, "F": 16},
        [[date_type(2020, 5, 1), None, "근무", "전일제", "Y"],
         [date_type(2024, 11, 1), None, "근무", "시간제", "N"]],
    )

    # 시트6: 필요인력
    ws6 = wb_t.create_sheet("필요인력")
    ws6["A1"] = "필요인력 확보 현황 (별도 보상제)"
    ws6["A1"].font = TITLE_FONT
    ws6.merge_cells("A1:B1")
    ws6["A3"] = "※ Y 또는 N만 입력하세요."
    ws6["A3"].font = NOTE_FONT
    ws6.merge_cells("A3:B3")
    nec_items = [
        ("약사 상근 여부", "Y"),
        ("약사 주16시간 이상 근무 (환자 200명 미만 시 인정)", "N"),
        ("보건의료정보관리사 상근 1인 이상", "Y"),
        ("방사선사 상근 1인 이상", "Y"),
        ("임상병리사 상근 1인 이상", "Y"),
        ("물리치료사 상근 1인 이상", "Y"),
        ("사회복지사 상근 1인 이상", "Y"),
    ]
    style_header(ws6, ["A4", "B4"])
    ws6["A4"] = "항목"; ws6["B4"] = "Y/N"
    for r, (label, val) in enumerate(nec_items, start=5):
        ws6[f"A{r}"] = label
        ws6[f"A{r}"].font = NORMAL_FONT
        ws6[f"A{r}"].border = BORDER
        ws6[f"B{r}"] = val
        style_input(ws6, f"B{r}:B{r}")
    set_widths(ws6, {"A": 46, "B": 10})

    buf = io.BytesIO()
    wb_t.save(buf)
    return buf.getvalue()

# ──────────────────────────────────────────────
# 엑셀 업로드 — 세션 key에 직접 쓰기
# ──────────────────────────────────────────────
QUARTER_KEYS = list(QUARTER_RANGES.keys())

def apply_uploaded_data(parsed):
    st.session_state["hosp_name"] = parsed["hosp_name"]
    st.session_state["year"] = parsed["year"]
    st.session_state["remote_area"] = parsed["remote_area"]

    q = parsed["quarter"]
    quarter = q if q in QUARTER_KEYS else QUARTER_KEYS[1]
    st.session_state["quarter_idx"] = QUARTER_KEYS.index(quarter)
    st.session_state["quarter_sel"] = quarter

    for i in range(3):
        for k in PAYER_KEYS:
            st.session_state[f"day_{i}_{k}"] = parsed["patients"][i][k]
            st.session_state[f"pat_{i}_{k}"] = parsed["daycare"][i][k]
        st.session_state.pop(f"month_editor_{i}", None)

    widget_prefixes = (
        "nu_hire_", "nu_resign_", "nu_status_", "nu_wt_",
        "ai_hire_", "ai_resign_", "ai_status_", "ai_wt_", "ai_employ_",
        "dr_hire_", "dr_resign_", "dr_status_", "dr_wt_", "dr_sp_",
    )
    for key in list(st.session_state):
        if key.startswith(widget_prefixes):
            del st.session_state[key]

    nurses = parsed["nurses"] or [{"hire_date": None, "resign_date": None,
                                    "status": "근무", "worktype": WORKTYPE_OPTS[0]}]
    aides = parsed["aides"] or [{"hire_date": None, "resign_date": None, "status": "근무",
                                  "worktype": WORKTYPE_OPTS[0], "employ": "정규직"}]
    doctors = parsed["doctors"] or [{"hire_date": None, "resign_date": None,
                                      "status": "근무", "worktype": "전일제", "specialist": True}]

    st.session_state.nurse_rows = nurses
    st.session_state.aide_rows = aides
    st.session_state.doctor_rows = doctors

    for i, n in enumerate(nurses):
        st.session_state[f"nu_hire_{i}"] = n["hire_date"]
        st.session_state[f"nu_resign_{i}"] = n["resign_date"]
        st.session_state[f"nu_status_{i}"] = n["status"]
        st.session_state[f"nu_wt_{i}"] = n["worktype"]

    for i, a in enumerate(aides):
        st.session_state[f"ai_hire_{i}"] = a["hire_date"]
        st.session_state[f"ai_resign_{i}"] = a["resign_date"]
        st.session_state[f"ai_status_{i}"] = a["status"]
        st.session_state[f"ai_wt_{i}"] = a["worktype"]
        st.session_state[f"ai_employ_{i}"] = a["employ"]

    for i, d in enumerate(doctors):
        st.session_state[f"dr_hire_{i}"] = d["hire_date"]
        st.session_state[f"dr_resign_{i}"] = d["resign_date"]
        st.session_state[f"dr_status_{i}"] = d["status"]
        st.session_state[f"dr_wt_{i}"] = d["worktype"]
        st.session_state[f"dr_sp_{i}"] = d["specialist"]

    for k, v in parsed["necessary"].items():
        st.session_state[k] = v

with st.expander("📂 엑셀 파일로 데이터 자동 입력 (클릭하여 열기)", expanded=False):
    st.markdown(
        "<div style='background:#e3f2fd;border:1px solid #90caf9;border-radius:8px;"
        "padding:12px 16px;font-size:13px;'><b>사용 방법</b><br>"
        "1. 아래 버튼으로 <b>입력 양식 엑셀 파일</b>을 다운로드하세요.<br>"
        "2. 양식의 <span style='background:#FFF9C4;padding:1px 4px;border-radius:3px;'>"
        "노란색 셀</span>에 기본정보·환자수현황·간호사·간호조무사·의사·필요인력을 입력 후 저장하세요.<br>"
        "3. 저장한 파일을 아래 업로드 칸에 올리면 모든 항목이 자동으로 채워집니다.</div>",
        unsafe_allow_html=True
    )
    template_bytes = generate_template_bytes()
    st.download_button(
        label="⬇️ 입력 양식 다운로드 (Excel)", data=template_bytes,
        file_name="요양병원_데이터입력양식.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    uploaded = st.file_uploader("작성한 엑셀 파일 업로드", type=["xlsx"], key="excel_upload")
    if uploaded is not None:
        uploaded_bytes = uploaded.getvalue()
        file_signature = hashlib.sha256(uploaded_bytes).hexdigest()
        reapply = st.button("업로드한 데이터 다시 적용", key="reapply_excel")
        if st.session_state.get("_applied_excel_signature") != file_signature or reapply:
            parsed, err = parse_excel_upload(uploaded_bytes)
            if err:
                st.error("파싱 오류: " + err)
            elif parsed:
                apply_uploaded_data(parsed)
                st.session_state["_applied_excel_signature"] = file_signature
                st.session_state["_excel_upload_message"] = (
                    "데이터 로드 완료! 간호사 " + str(len(parsed["nurses"])) +
                    "명 / 간호조무사 " + str(len(parsed["aides"])) +
                    "명 / 의사 " + str(len(parsed["doctors"])) + "명 입력됨."
                )
                st.rerun()
        elif "_excel_upload_message" in st.session_state:
            st.success(st.session_state.pop("_excel_upload_message"))

# ──────────────────────────────────────────────
# ① 기본 정보
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">① 기본 정보</div>', unsafe_allow_html=True)
col1, col2, col3, col4 = st.columns([2, 1, 1.4, 1.4])
with col1:
    hosp_name = st.text_input("요양기관명", key="hosp_name", placeholder="예: OO요양병원")
with col2:
    year = st.number_input("연도", min_value=2020, max_value=2040, step=1, key="year")
with col3:
    quarter_label = st.selectbox("적용 분기", QUARTER_KEYS, index=st.session_state["quarter_idx"], key="quarter_sel")
with col4:
    remote_area = st.checkbox("의료취약지역 소재 요양기관", key="remote_area",
                               help="「소득세법 시행규칙」제7조제4호에 따른 의료취약지역 소재 기관은 단시간 근무 간호인력 산정 가중치가 상향 적용됩니다.")

q_start, q_end, q_days = get_quarter_dates(year, quarter_label)
st.info(
    f"📅 산정대상 기간(전전분기 말월 15일 ~ 전분기 말월 14일): **{q_start}** ~ **{q_end}**  |  총 **{q_days}일**  "
    f"→ 산정된 등급은 익분기(다음 분기) 요양병원입원료 등에 적용됩니다."
)

# ──────────────────────────────────────────────
# ② 환자 수 현황
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">② 환자 수 현황 (재원환자수 · 낮병동환자수)</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="yellow-note">🟡 각 월별 재원환자수(환자별 재원일수의 합)와 낮병동 입원환자수를 재원형태(건강보험/의료급여/자보/산재/기타)별로 입력하세요. '
    '낮병동 입원환자 1인은 입원환자 1인으로 환산됩니다.</div>', unsafe_allow_html=True
)

total_inpatient = 0
total_daycare = 0
month_cols = st.columns(3)
for i in range(3):
    lbl = month_label(q_start, i)
    with month_cols[i]:
        st.markdown(f'<div class="sub-title">{lbl}</div>', unsafe_allow_html=True)
        df_default = pd.DataFrame(
            [[st.session_state[f"day_{i}_{k}"] for k in PAYER_KEYS],
             [st.session_state[f"pat_{i}_{k}"] for k in PAYER_KEYS]],
            index=["입원환자수(재원일수합)", "낮병동환자수"], columns=PAYER_COLS
        )
        edited = st.data_editor(
            df_default, key=f"month_editor_{i}", width='stretch',
            num_rows="fixed"
        )
        for j, k in enumerate(PAYER_KEYS):
            st.session_state[f"day_{i}_{k}"] = int(edited.iloc[0, j])
            st.session_state[f"pat_{i}_{k}"] = int(edited.iloc[1, j])
        month_in = int(edited.iloc[0].sum())
        month_day = int(edited.iloc[1].sum())
        st.caption(f"입원환자수 합계 {month_in}명 · 낮병동 합계 {month_day}명")
        total_inpatient += month_in
        total_daycare += month_day

avg_patients = (total_inpatient + total_daycare) / q_days if q_days > 0 else 0
st.markdown(
    f'<div class="result-card">📊 <b>적용입원환자수 (3개월 평균)</b> = '
    f'({total_inpatient}명 + {total_daycare}명) ÷ {q_days}일 = '
    f'<span class="kpi-value">{avg_patients:.2f}명</span></div>', unsafe_allow_html=True
)

# ──────────────────────────────────────────────
# ③ 간호사 인력
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">③ 간호사 인력 현황</div>', unsafe_allow_html=True)
ptable = PARTTIME_TABLE_REMOTE if remote_area else PARTTIME_TABLE_NORMAL
st.markdown(
    '<div class="yellow-note">🟡 단시간 근무 가중치 자동 적용 '
    + (" (의료취약지역 상향기준)" if remote_area else "")
    + ": " + " / ".join([f"{k.split('(')[1][:-1]}={v}" for k, v in ptable.items()]) + "</div>",
    unsafe_allow_html=True
)

c1, c2, _ = st.columns([2.8, 3.0, 14])
with c1:
    if st.button("➕ 간호사 추가", width='stretch'):
        st.session_state.nurse_rows.append({"hire_date": None, "resign_date": None,
                                             "status": "근무", "worktype": WORKTYPE_OPTS[0]})
with c2:
    if st.button("➖ 마지막 행 삭제", key="del_nurse", width='stretch') and len(st.session_state.nurse_rows) > 1:
        st.session_state.nurse_rows.pop()

hc = st.columns([0.4, 1.5, 1.5, 1.3, 2.2, 1.4, 1.4])
for c, t in zip(hc, ["#", "입사일", "퇴사일", "상태", "근무형태", "산정일수 🟡", "환산인원 🟡"]):
    c.markdown(f"**{t}**")

nurse_total = 0.0
for i, row in enumerate(st.session_state.nurse_rows):
    cols = st.columns([0.4, 1.5, 1.5, 1.3, 2.2, 1.4, 1.4])
    cols[0].markdown(f"{i+1}")
    hire = cols[1].date_input("입사일", value=row["hire_date"], key=f"nu_hire_{i}", label_visibility="collapsed")
    resign_disabled = (row["status"] != "퇴사")
    resign = cols[2].date_input("퇴사일", value=row["resign_date"], key=f"nu_resign_{i}",
                                 label_visibility="collapsed", disabled=resign_disabled)
    status = cols[3].selectbox("상태", ["근무", "퇴사"], index=0 if row["status"] == "근무" else 1,
                                key=f"nu_status_{i}", label_visibility="collapsed")
    worktype = cols[4].selectbox("근무형태", WORKTYPE_OPTS,
                                  index=WORKTYPE_OPTS.index(row["worktype"]) if row["worktype"] in WORKTYPE_OPTS else 0,
                                  key=f"nu_wt_{i}", label_visibility="collapsed")
    resign_final = resign if status == "퇴사" else None
    days = calc_active_days(hire, status, q_start, q_end, resign_final)
    weight = ptable[worktype]
    fte = (days / q_days) * weight if q_days > 0 else 0.0
    cols[5].markdown(f'<div style="padding-top:8px;color:#0d4f3c;font-weight:600">{days}일</div>', unsafe_allow_html=True)
    cols[6].markdown(f'<div style="padding-top:8px;color:#0d4f3c;font-weight:600">{fte:.3f}명</div>', unsafe_allow_html=True)
    nurse_total += fte
    st.session_state.nurse_rows[i] = {"hire_date": hire, "resign_date": resign_final,
                                       "status": status, "worktype": worktype}

st.markdown(
    f'<div class="result-card">👩‍⚕️ <b>간호사 수 (3개월 평균, ③)</b>: '
    f'<span class="kpi-value">{nurse_total:.2f}명</span></div>', unsafe_allow_html=True
)

# ──────────────────────────────────────────────
# ④ 간호조무사 인력
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">④ 간호조무사 인력 현황</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="yellow-note">🟡 계약직 간호조무사가 전일제로 근무하는 경우, 3인을 2인으로 산정합니다(환산계수 0.6667). '
    '계약직 간호인력의 정규직 의무고용비율은 80% 이상이어야 등급이 정상 인정됩니다.</div>', unsafe_allow_html=True
)

c3, c4, _ = st.columns([3.2, 3.0, 14])
with c3:
    if st.button("➕ 간호조무사 추가", width='stretch'):
        st.session_state.aide_rows.append({"hire_date": None, "resign_date": None,
                                            "status": "근무", "worktype": WORKTYPE_OPTS[0], "employ": "정규직"})
with c4:
    if st.button("➖ 마지막 행 삭제 ", key="del_aide", width='stretch') and len(st.session_state.aide_rows) > 1:
        st.session_state.aide_rows.pop()

hc2 = st.columns([0.4, 1.4, 1.4, 1.1, 1.9, 1.1, 1.3, 1.3])
for c, t in zip(hc2, ["#", "입사일", "퇴사일", "상태", "근무형태", "고용형태", "산정일수 🟡", "환산인원 🟡"]):
    c.markdown(f"**{t}**")

aide_total = 0.0
for i, row in enumerate(st.session_state.aide_rows):
    cols = st.columns([0.4, 1.4, 1.4, 1.1, 1.9, 1.1, 1.3, 1.3])
    cols[0].markdown(f"{i+1}")
    hire = cols[1].date_input("입사일", value=row["hire_date"], key=f"ai_hire_{i}", label_visibility="collapsed")
    resign_disabled = (row["status"] != "퇴사")
    resign = cols[2].date_input("퇴사일", value=row["resign_date"], key=f"ai_resign_{i}",
                                 label_visibility="collapsed", disabled=resign_disabled)
    status = cols[3].selectbox("상태", ["근무", "퇴사"], index=0 if row["status"] == "근무" else 1,
                                key=f"ai_status_{i}", label_visibility="collapsed")
    worktype = cols[4].selectbox("근무형태", WORKTYPE_OPTS,
                                  index=WORKTYPE_OPTS.index(row["worktype"]) if row["worktype"] in WORKTYPE_OPTS else 0,
                                  key=f"ai_wt_{i}", label_visibility="collapsed")
    employ = cols[5].selectbox("고용형태", ["정규직", "계약직"],
                                index=0 if row.get("employ", "정규직") == "정규직" else 1,
                                key=f"ai_employ_{i}", label_visibility="collapsed")
    resign_final = resign if status == "퇴사" else None
    days = calc_active_days(hire, status, q_start, q_end, resign_final)
    weight = ptable[worktype]
    fte = (days / q_days) * weight if q_days > 0 else 0.0
    if employ == "계약직" and worktype == WORKTYPE_OPTS[0]:
        fte *= 0.6667
    cols[6].markdown(f'<div style="padding-top:8px;color:#0d4f3c;font-weight:600">{days}일</div>', unsafe_allow_html=True)
    cols[7].markdown(f'<div style="padding-top:8px;color:#0d4f3c;font-weight:600">{fte:.3f}명</div>', unsafe_allow_html=True)
    aide_total += fte
    st.session_state.aide_rows[i] = {"hire_date": hire, "resign_date": resign_final,
                                      "status": status, "worktype": worktype, "employ": employ}

st.markdown(
    f'<div class="result-card">🧑‍🔧 <b>간호조무사 수 (3개월 평균, ④)</b>: '
    f'<span class="kpi-value">{aide_total:.2f}명</span></div>', unsafe_allow_html=True
)

nursing_personnel_total = nurse_total + aide_total
patient_to_nursing_ratio = (avg_patients / nursing_personnel_total) if nursing_personnel_total > 0 else 999
patient_to_nurse_ratio = (avg_patients / nurse_total) if nurse_total > 0 else 999
nurse_share = (nurse_total / nursing_personnel_total) if nursing_personnel_total > 0 else 0

nursing_grade = determine_nursing_grade(patient_to_nursing_ratio)
over18_penalty = (patient_to_nurse_ratio > 18) and (nursing_grade <= 5)
nurse_ratio_bonus = (nurse_share >= (2/3)) and (nursing_grade <= 5)

# ──────────────────────────────────────────────
# ⑤ 의사 인력
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">⑤ 의사 인력 현황</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="yellow-note">🟡 시간제·격일제 의사는 주3일 이상이면서 주20시간 이상 근무 시 0.5인으로 산정합니다. '
    '전문의 여부는 「전문의의 수련 및 자격 인정 등에 관한 규정」상 26개 전문과목 인정 여부를 기준으로 체크하세요.</div>',
    unsafe_allow_html=True
)

c5, c6, _ = st.columns([2.6, 3.0, 14])
with c5:
    if st.button("➕ 의사 추가", width='stretch'):
        st.session_state.doctor_rows.append({"hire_date": None, "resign_date": None,
                                              "status": "근무", "worktype": "전일제", "specialist": True})
with c6:
    if st.button("➖ 마지막 행 삭제  ", key="del_doc", width='stretch') and len(st.session_state.doctor_rows) > 1:
        st.session_state.doctor_rows.pop()

hc3 = st.columns([0.4, 1.5, 1.5, 1.2, 1.8, 1.3, 1.3, 1.3])
for c, t in zip(hc3, ["#", "입사일", "퇴사일", "상태", "근무형태", "전문의", "산정일수 🟡", "환산인원 🟡"]):
    c.markdown(f"**{t}**")

doctor_total = 0.0
specialist_total = 0.0
for i, row in enumerate(st.session_state.doctor_rows):
    cols = st.columns([0.4, 1.5, 1.5, 1.2, 1.8, 1.3, 1.3, 1.3])
    cols[0].markdown(f"{i+1}")
    hire = cols[1].date_input("입사일", value=row["hire_date"], key=f"dr_hire_{i}", label_visibility="collapsed")
    resign_disabled = (row["status"] != "퇴사")
    resign = cols[2].date_input("퇴사일", value=row["resign_date"], key=f"dr_resign_{i}",
                                 label_visibility="collapsed", disabled=resign_disabled)
    status = cols[3].selectbox("상태", ["근무", "퇴사"], index=0 if row["status"] == "근무" else 1,
                                key=f"dr_status_{i}", label_visibility="collapsed")
    worktype = cols[4].selectbox("근무형태", ["전일제", "시간제/격일제(0.5인)"],
                                  index=0 if row.get("worktype", "전일제") == "전일제" else 1,
                                  key=f"dr_wt_{i}", label_visibility="collapsed")
    specialist = cols[5].checkbox("전문의", value=row.get("specialist", True), key=f"dr_sp_{i}",
                                   label_visibility="collapsed")
    resign_final = resign if status == "퇴사" else None
    days = calc_active_days(hire, status, q_start, q_end, resign_final)
    base_fte = (days / q_days) if q_days > 0 else 0.0
    fte = base_fte * (0.5 if worktype != "전일제" else 1.0)
    cols[6].markdown(f'<div style="padding-top:8px;color:#0d4f3c;font-weight:600">{days}일</div>', unsafe_allow_html=True)
    cols[7].markdown(f'<div style="padding-top:8px;color:#0d4f3c;font-weight:600">{fte:.3f}명</div>', unsafe_allow_html=True)
    doctor_total += fte
    if specialist:
        specialist_total += fte
    st.session_state.doctor_rows[i] = {"hire_date": hire, "resign_date": resign_final,
                                        "status": status, "worktype": worktype, "specialist": specialist}

st.markdown(
    f'<div class="result-card">🩺 <b>의사 수 (3개월 평균, ⑩)</b>: <span class="kpi-value">{doctor_total:.2f}명</span>'
    f'&nbsp;&nbsp;|&nbsp;&nbsp;🎖️ <b>전문의 수 (⑫)</b>: <span class="kpi-value">{specialist_total:.2f}명</span></div>',
    unsafe_allow_html=True
)

patient_to_doctor_ratio = (avg_patients / doctor_total) if doctor_total > 0 else 999
specialist_ratio = (specialist_total / doctor_total) if doctor_total > 0 else 0
doctor_grade = determine_doctor_grade(patient_to_doctor_ratio, specialist_ratio)

# ──────────────────────────────────────────────
# ⑥ 필요인력 확보 현황
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">⑥ 필요인력 확보 현황 (별도 보상제)</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="yellow-note">🟡 약사가 상근(적용입원환자수 200명 이상) 또는 주16시간 이상 근무(200명 미만)이면서, '
    '보건의료정보관리사·방사선사·임상병리사·물리치료사·사회복지사 중 상근 1명 이상인 직종이 4개 이상이면 '
    '1일당 1,710원을 별도 산정합니다. (해당 치료실·장비를 실제 보유·사용하는 경우에 한함)</div>', unsafe_allow_html=True
)

nc1, nc2, nc3 = st.columns(3)
with nc1:
    pharm_present = st.checkbox("약사 상근", key="pharm_present")
    pharm_hours16 = st.checkbox("약사 주16시간↑ 근무 (환자 200명 미만 시 인정)", key="pharm_hours16")
with nc2:
    hima_present = st.checkbox("보건의료정보관리사 상근 1인 이상", key="hima_present")
    radio_present = st.checkbox("방사선사 상근 1인 이상", key="radio_present")
with nc3:
    lab_present = st.checkbox("임상병리사 상근 1인 이상", key="lab_present")
    pt_present = st.checkbox("물리치료사 상근 1인 이상", key="pt_present")
sw_present = st.checkbox("사회복지사 상근 1인 이상", key="sw_present")

pharm_ok = pharm_present or (avg_patients < 200 and pharm_hours16)
other_flags = [hima_present, radio_present, lab_present, pt_present, sw_present]
other_count = sum(other_flags)
necessary_staff_bonus = pharm_ok and (other_count >= 4)

st.markdown(
    f'<div class="result-card">🧾 필요인력 요건 충족 직종 수: <b>{other_count} / 5</b> (+약사 조건: '
    f'{"충족" if pharm_ok else "미충족"}) → 별도 보상 {"✅ 적용 (1일당 1,710원)" if necessary_staff_bonus else "❌ 미적용"}</div>',
    unsafe_allow_html=True
)

# ──────────────────────────────────────────────
# ⑦ 등급 산정 결과 보고서
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">⑦ 등급 산정 결과 보고서</div>', unsafe_allow_html=True)

k1, k2, k3, k4, k5, k6 = st.columns(6)
def kpi(col, label, value, unit=""):
    col.markdown(
        f'<div class="result-card" style="text-align:center">'
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{value}</div>'
        f'<div class="kpi-unit">{unit}</div></div>', unsafe_allow_html=True
    )

kpi(k1, "👥 적용입원환자수", f"{avg_patients:.2f}", "명 (3개월평균)")
kpi(k2, "🧮 환자수 대 간호인력수", f"{patient_to_nursing_ratio:.2f} : 1", "②/(③+④)")
kpi(k3, "🧮 환자수 대 간호사수", f"{patient_to_nurse_ratio:.2f} : 1", "②/③")
kpi(k4, "🧮 간호사수 대 간호인력수", f"{nurse_share*100:.1f}%", "③/(③+④)")
kpi(k5, "🧮 환자수 대 의사수", f"{patient_to_doctor_ratio:.2f} : 1", "⑨/⑩")
kpi(k6, "🎖️ 전문의 비율", f"{specialist_ratio*100:.1f}%", "⑫/⑩")

st.markdown("---")
gcol1, gcol2 = st.columns(2)
with gcol1:
    st.markdown(f"""
    <div style="text-align:center; margin:14px 0;">
        <div style="font-size:15px; color:#555; margin-bottom:6px;">간호인력 확보수준 등급</div>
        <span class="grade-box {grade_css_n(nursing_grade)}">{nursing_grade}등급</span>
        <div style="font-size:13px; color:#777; margin-top:8px;">기본 가산율 <b>{NURSING_BONUS_PCT[nursing_grade]:+d}%</b>
        {" · 18:1 초과 추가감산 -15% 적용" if over18_penalty else ""}
        {" · 간호사비율 2/3↑ 1일 2,000원 가산" if nurse_ratio_bonus else ""}</div>
    </div>""", unsafe_allow_html=True)
with gcol2:
    st.markdown(f"""
    <div style="text-align:center; margin:14px 0;">
        <div style="font-size:15px; color:#555; margin-bottom:6px;">의사인력 확보수준 등급</div>
        <span class="grade-box {grade_css_d(doctor_grade)}">{doctor_grade}등급</span>
        <div style="font-size:13px; color:#777; margin-top:8px;">가산율 <b>{DOCTOR_BONUS_PCT[doctor_grade]:+d}%</b></div>
    </div>""", unsafe_allow_html=True)

st.markdown("---")
rc1, rc2 = st.columns(2)
with rc1:
    st.markdown("##### 📋 간호인력 등급 기준표")
    ndf = pd.DataFrame(NURSING_GRADE_TABLE, columns=["등급", "환자수 대 간호인력수 기준", "가산율"])
    ndf["현재"] = ["✅" if g == nursing_grade else "" for g, _, _ in NURSING_GRADE_TABLE]
    st.table(ndf.set_index("등급"))
with rc2:
    st.markdown("##### 📋 의사인력 등급 기준표")
    ddf = pd.DataFrame(DOCTOR_GRADE_TABLE, columns=["등급", "환자수 대 의사수 및 전문의비율 기준", "가산율"])
    ddf["현재"] = ["✅" if g == doctor_grade else "" for g, _, _ in DOCTOR_GRADE_TABLE]
    st.table(ddf.set_index("등급"))

# 예시 금액 계산 (참고용)
BASE_SCORE = 270.04     # 요양병원입원료(요-51) 소정점수 예시값(의료경도 기준)
UNIT_PRICE = 84.2       # 2025년 기준 요양병원 점수당 단가(원) - 매년 변경되므로 참고용
base_amount = BASE_SCORE * UNIT_PRICE
nursing_pct_effective = NURSING_BONUS_PCT[nursing_grade] - (15 if over18_penalty else 0)
nursing_addon = base_amount * nursing_pct_effective / 100
doctor_addon = base_amount * DOCTOR_BONUS_PCT[doctor_grade] / 100
flat_bonus = (2000 if nurse_ratio_bonus else 0) + (1710 if necessary_staff_bonus else 0)
est_daily_total = base_amount + nursing_addon + doctor_addon + flat_bonus

st.markdown("---")
st.markdown("##### 💰 1일당 요양병원입원료 예시 계산 (참고용)")
st.caption("※ 실제 수가는 환자군(의료최고도~선택입원군)별 상대가치점수 및 당해연도 점수당 단가에 따라 달라집니다. "
           "아래는 대표 점수(270.04점, 의료경도 예시)를 기준으로 한 참고 추정치이며, 정확한 청구금액은 심사평가원 고시를 확인하십시오.")
calc_df = pd.DataFrame({
    "항목": ["요양병원입원료 기준금액(예시)", f"간호인력등급 가산/감산 ({nursing_pct_effective:+d}%)",
             f"의사인력등급 가산/감산 ({DOCTOR_BONUS_PCT[doctor_grade]:+d}%)",
             "간호사비율 2/3↑ 가산", "필요인력 확보 가산", "1일당 추정 합계"],
    "금액(원)": [f"{base_amount:,.0f}", f"{nursing_addon:+,.0f}", f"{doctor_addon:+,.0f}",
               f"{2000 if nurse_ratio_bonus else 0:+,.0f}", f"{1710 if necessary_staff_bonus else 0:+,.0f}",
               f"{est_daily_total:,.0f}"]
})
st.table(calc_df.set_index("항목"))

with st.expander("🔍 상세 계산 내역 보기"):
    st.markdown(f"""
| 항목 | 계산식 | 결과 |
|------|--------|------|
| 산정 기간 | {q_start} ~ {q_end} | **{q_days}일** |
| 총 재원환자수(입원+낮병동) | {total_inpatient}명 + {total_daycare}명 | |
| 적용입원환자수(②) | ({total_inpatient}+{total_daycare}) ÷ {q_days} | **{avg_patients:.2f}명** |
| 간호사 수(③) | 근무일수 ÷ {q_days} × 근무형태가중치 합계 | **{nurse_total:.2f}명** |
| 간호조무사 수(④) | 근무일수 ÷ {q_days} × 근무형태가중치(계약직 보정 포함) 합계 | **{aide_total:.2f}명** |
| 환자수 대 간호인력수 | {avg_patients:.2f} ÷ ({nurse_total:.2f}+{aide_total:.2f}) | **{patient_to_nursing_ratio:.2f} : 1** |
| 환자수 대 간호사수 | {avg_patients:.2f} ÷ {nurse_total:.2f} | **{patient_to_nurse_ratio:.2f} : 1** (18:1 초과 시 추가 -15%) |
| 간호사수 대 간호인력수 | {nurse_total:.2f} ÷ ({nurse_total:.2f}+{aide_total:.2f}) | **{nurse_share*100:.1f}%** (2/3 이상 시 1일 2,000원 가산) |
| 의사 수(⑩) | 근무일수 ÷ {q_days} × (전일제 1.0 / 시간제 0.5) | **{doctor_total:.2f}명** |
| 전문의 수(⑫) | 전문의 해당 인력 환산 합계 | **{specialist_total:.2f}명** |
| 환자수 대 의사수 | {avg_patients:.2f} ÷ {doctor_total:.2f} | **{patient_to_doctor_ratio:.2f} : 1** |
| 전문의 비율 | {specialist_total:.2f} ÷ {doctor_total:.2f} | **{specialist_ratio*100:.1f}%** |
| **간호인력 등급** | | **{nursing_grade}등급 ({NURSING_BONUS_PCT[nursing_grade]:+d}%)** |
| **의사인력 등급** | | **{doctor_grade}등급 ({DOCTOR_BONUS_PCT[doctor_grade]:+d}%)** |
""")

# ──────────────────────────────────────────────
# ⑧ 등급 상향 시뮬레이션
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">⑧ 등급 상향 인력 충원 시뮬레이션 (자동 계산)</div>', unsafe_allow_html=True)

NURSING_UPPER = {1: 0, 2: 4.5, 3: 5.0, 4: 5.5, 5: 6.0, 6: 6.5}
sim_rows = []
for g in range(1, nursing_grade):
    target_ratio = NURSING_UPPER[g + 1] - 0.01 if g < 5 else NURSING_UPPER[g + 1] - 0.01
    target_ratio = NURSING_UPPER[g + 1] - 0.01
    needed_total = avg_patients / target_ratio if target_ratio > 0 else 0
    additional = max(0, needed_total - nursing_personnel_total)
    sim_rows.append({
        "목표 간호등급": f"{g}등급", "필요 총 간호인력(환산)": f"{needed_total:.2f}명",
        "추가 필요 인력(환산)": f"{additional:.2f}명",
        "전일제 충원 시": f"{additional:.2f}명 추가",
        "주36~40h 단시간 충원 시": f"{additional/0.9:.2f}명 추가",
    })
if sim_rows:
    st.markdown("##### 간호인력 등급 상향")
    st.dataframe(pd.DataFrame(sim_rows), width='stretch', hide_index=True)
else:
    st.success("🎉 간호인력은 이미 최고 등급(1등급)입니다!")

doc_sim_rows = []
DOCTOR_UPPER = {1: 35, 2: 35, 3: 40, 4: 999}
if doctor_grade == 4:
    needed = avg_patients / 40
    doc_sim_rows.append({"목표 의사등급": "3등급", "필요 총 의사수(환산)": f"{needed:.2f}명",
                          "추가 필요 의사수": f"{max(0, needed - doctor_total):.2f}명"})
if doctor_grade >= 3:
    needed = avg_patients / 35
    doc_sim_rows.append({"목표 의사등급": "1~2등급", "필요 총 의사수(환산)": f"{needed:.2f}명",
                          "추가 필요 의사수": f"{max(0, needed - doctor_total):.2f}명"})
if doctor_grade == 2:
    doc_sim_rows.append({"목표 의사등급": "1등급 (전문의비율 50%↑ 필요)",
                          "필요 총 의사수(환산)": f"{doctor_total:.2f}명 (동일)",
                          "추가 필요 의사수": f"전문의 비율 {(0.5*doctor_total - specialist_total):.2f}명분 상향 필요"})
if doc_sim_rows:
    st.markdown("##### 의사인력 등급 상향")
    st.dataframe(pd.DataFrame(doc_sim_rows), width='stretch', hide_index=True)
else:
    st.success("🎉 의사인력은 이미 최고 등급(1등급)입니다!")

# ──────────────────────────────────────────────
# ⑨ AI 등급 진단 및 컨설팅 보고서
# ──────────────────────────────────────────────
from google import genai
from google.genai import types
import json
import os

GEMINI_MODEL_PRIMARY = "gemini-3.6-flash"
GEMINI_MODEL_FALLBACK = "gemini-3.5-flash-lite"

def _get_gemini_api_key():
    try:
        if "GEMINI_API_KEY" in st.secrets:
            return st.secrets["GEMINI_API_KEY"]
    except Exception:
        pass
    return os.environ.get("GEMINI_API_KEY")

st.markdown('<div class="section-title">⑨ AI 등급 진단 및 컨설팅 보고서</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="yellow-note">🤖 현재 입력된 간호·의사·필요인력 데이터를 바탕으로 AI가 등급 현황을 종합 진단하고, '
    '등급 상향 및 필요인력 보상 확보를 위한 구체적인 인력 충원·운영 전략을 제안합니다.</div>',
    unsafe_allow_html=True
)

analysis_data = {
    "요양기관명": hosp_name or "(미입력)", "적용분기": quarter_label, "연도": year,
    "산정기간": f"{q_start} ~ {q_end}", "산정일수": q_days,
    "적용입원환자수_3개월평균": round(avg_patients, 2),
    "간호사수_3개월평균": round(nurse_total, 2),
    "간호조무사수_3개월평균": round(aide_total, 2),
    "환자수_대_간호인력수": round(patient_to_nursing_ratio, 2),
    "환자수_대_간호사수": round(patient_to_nurse_ratio, 2),
    "간호사비율": round(nurse_share * 100, 1),
    "간호등급": nursing_grade, "간호등급_가산율": NURSING_BONUS_PCT[nursing_grade],
    "18대1초과_추가감산_적용여부": over18_penalty,
    "간호사비율2/3가산_적용여부": nurse_ratio_bonus,
    "의사수_3개월평균": round(doctor_total, 2), "전문의수_3개월평균": round(specialist_total, 2),
    "환자수_대_의사수": round(patient_to_doctor_ratio, 2), "전문의비율": round(specialist_ratio * 100, 1),
    "의사등급": doctor_grade, "의사등급_가산율": DOCTOR_BONUS_PCT[doctor_grade],
    "필요인력_요건충족직종수": other_count, "필요인력_보상적용여부": necessary_staff_bonus,
    "1일당_추정_요양병원입원료_참고값": round(est_daily_total, 0),
}

if st.button("🤖 AI 컨설팅 보고서 생성", type="primary", width='stretch'):
    api_key = _get_gemini_api_key()
    if not api_key:
        st.error("AI 분석 오류: GEMINI_API_KEY가 설정되어 있지 않습니다.")
        st.info("💡 Streamlit Cloud의 Settings → Secrets에 `GEMINI_API_KEY = \"발급받은키\"` 형식으로 등록해주세요.")
    else:
        with st.spinner("AI가 데이터를 분석하고 컨설팅 보고서를 작성 중입니다..."):
            system_prompt = """당신은 대한민국 요양병원 경영 및 수가 전문 컨설턴트입니다.
특히 요양병원 간호·의사 인력확보수준 입원료 차등제 및 필요인력 보상제 관리에 특화된 전문가로서,
'주식회사 메디엄'의 수석 컨설턴트입니다.
보고서는 전문적이고 수치 기반으로, 마크다운 형식으로 아래 구조를 따르세요:
# 요양병원 입원료 차등제 등급 진단 컨설팅 보고서
## 1. 현황 요약
## 2. 핵심 지표 분석 (간호인력 · 의사인력 · 필요인력)
## 3. 간호등급 상향 전략 (단계별 충원 시나리오)
### 시나리오 A: 간호사 우선 충원
### 시나리오 B: 간호조무사 우선 충원
### 시나리오 C: 혼합 충원 (최적안)
## 4. 의사등급 상향 전략 (전문의 비율 개선 포함)
## 5. 필요인력 별도보상제 확보 전략
## 6. 재정적 효과 분석 (등급 상향 시 1일당·연간 추가수익 추정)
## 7. 리스크 및 주의사항 (18:1 초과감산, 적정성평가 환류 리스크 등)
## 8. 종합 권고사항 및 실행 로드맵"""

            user_prompt = (
                f"다음 데이터를 분석해 전문 컨설팅 보고서를 작성해주세요.\n\n"
                f"```json\n{json.dumps(analysis_data, ensure_ascii=False, indent=2)}\n```\n\n"
                f"간호등급 {nursing_grade}등급, 의사등급 {doctor_grade}등급 각각에 대해 "
                f"한 단계 상향을 위한 구체적 인력 충원 시나리오(전일제/단시간 근무형태별)와, "
                f"필요인력 보상제 확보를 위한 실행 방안을 포함해주세요. "
                f"18:1 초과 여부와 간호사비율 2/3 가산 유지 전략도 함께 다뤄주세요."
            )

            client = genai.Client(api_key=api_key)
            report_placeholder = st.empty()
            full_report = ""
            last_error = None
            success = False

            for model_name in [GEMINI_MODEL_PRIMARY, GEMINI_MODEL_FALLBACK]:
                try:
                    stream = client.models.generate_content_stream(
                        model=model_name,
                        contents=user_prompt,
                        config=types.GenerateContentConfig(
                            system_instruction=system_prompt, max_output_tokens=4000
                        ),
                    )
                    full_report = ""
                    for chunk in stream:
                        if chunk.text:
                            full_report += chunk.text
                            report_placeholder.markdown(full_report)
                    if full_report.strip():
                        success = True
                        break
                except Exception as e:
                    last_error = e
                    full_report = ""
                    continue

            if success:
                st.session_state["last_report"] = full_report
                st.success("✅ AI 컨설팅 보고서 생성 완료!")
            else:
                st.error(f"AI 분석 오류: {str(last_error)}")
                st.info("💡 GEMINI_API_KEY 값과 모델 사용 가능 여부(할당량 포함)를 확인해주세요.")

elif "last_report" in st.session_state:
    st.markdown(st.session_state["last_report"])
    st.info("💡 데이터를 변경한 후 버튼을 다시 누르면 새 보고서가 생성됩니다.")

# ──────────────────────────────────────────────
# 하단 푸터
# ──────────────────────────────────────────────
st.markdown(
    '<div class="footer">요양병원 입원료 차등제 등급 산정 시스템<br>'
    '<b>제작: 조정윤 · 주식회사 메디엄</b><br>'
    '본 시스템의 산정 결과는 참고용이며, 실제 신고 및 청구는 반드시 건강보험심사평가원 최신 고시를 기준으로 확인하시기 바랍니다.</div>',
    unsafe_allow_html=True
)
